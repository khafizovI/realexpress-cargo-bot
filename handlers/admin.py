"""Admin handlers."""
import logging
import os
from functools import wraps
from tempfile import NamedTemporaryFile

from aiogram import Router, types
from aiogram.filters import Command, StateFilter
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup

from database import db
from keyboards import admin_menu, main_menu
from subscription import normalize_channel

router = Router()


# --- States ---
class AddTrackState(StatesGroup):
    waiting_for_flight = State()
    waiting_for_codes = State()


class DeleteTrackState(StatesGroup):
    waiting_for_flight = State()


class AboutEditState(StatesGroup):
    choosing = State()
    waiting_for_uz_text = State()
    waiting_for_ru_text = State()


class PriceEditState(StatesGroup):
    choosing = State()
    waiting_for_uz_text = State()
    waiting_for_ru_text = State()


class AddressEditState(StatesGroup):
    choosing = State()
    waiting_for_uz_text = State()
    waiting_for_ru_text = State()
    waiting_for_map_link = State()


class AddAdminState(StatesGroup):
    waiting_for_user = State()


class LinkChannelState(StatesGroup):
    waiting_for_channel = State()


class ImportTracksState(StatesGroup):
    waiting_for_file = State()
    waiting_for_flight = State()

class BroadcastState(StatesGroup):
    waiting_for_content = State()


# --- Helpers ---
def is_admin_message(message: types.Message) -> bool:
    return db.is_admin(message.from_user.id)


def is_admin_user(user_id: int) -> bool:
    return db.is_admin(user_id)


def admin_only(handler):
    @wraps(handler)
    async def wrapper(message: types.Message, *args, **kwargs):
        if not is_admin_message(message):
            lang = db.get_user_language(message.from_user.id)
            await message.answer("Siz admin emassiz." if lang == "uz" else "Вы не администратор.")
            return
        return await handler(message, *args, **kwargs)

    return wrapper


def admin_only_callback(handler):
    @wraps(handler)
    async def wrapper(callback: types.CallbackQuery, *args, **kwargs):
        if not is_admin_user(callback.from_user.id):
            lang = db.get_user_language(callback.from_user.id)
            await callback.answer("Siz admin emassiz." if lang == "uz" else "Вы не администратор.", show_alert=True)
            return
        return await handler(callback, *args, **kwargs)

    return wrapper


def text_has(message: types.Message, *parts: str) -> bool:
    if not message.text:
        return False
    # Normalize apostrophes so "e'lon" matches "elon"
    def _norm(value: str) -> str:
        return (
            value.lower()
            .replace("’", "'")
            .replace("ʼ", "'")
            .replace("‘", "'")
            .replace("'", "")
        )

    text = _norm(message.text)
    return all(_norm(part) in text for part in parts)


def is_back_text(text: str) -> bool:
    return text in ("⬅️ Orqaga", "⬅️ Назад")


def price_edit_keyboard() -> types.InlineKeyboardMarkup:
    return types.InlineKeyboardMarkup(
        inline_keyboard=[
            [
                types.InlineKeyboardButton(text="✏️ O'zbekcha narxni tahrirlash", callback_data="edit_prices_uz"),
                types.InlineKeyboardButton(text="✏️ Русские цены", callback_data="edit_prices_ru"),
            ],
            [types.InlineKeyboardButton(text="⬅️ Orqaga", callback_data="admin_back")],
        ]
    )


def about_edit_keyboard() -> types.InlineKeyboardMarkup:
    return types.InlineKeyboardMarkup(
        inline_keyboard=[
            [
                types.InlineKeyboardButton(text="✏️ O'zbekcha matn", callback_data="edit_about_uz"),
                types.InlineKeyboardButton(text="✏️ Русский текст", callback_data="edit_about_ru"),
            ],
            [types.InlineKeyboardButton(text="⬅️ Orqaga", callback_data="admin_back")],
        ]
    )


def address_edit_keyboard() -> types.InlineKeyboardMarkup:
    return types.InlineKeyboardMarkup(
        inline_keyboard=[
            [
                types.InlineKeyboardButton(text="✏️ O'zbekcha manzil matni", callback_data="edit_address_uz"),
                types.InlineKeyboardButton(text="✏️ Русский адрес", callback_data="edit_address_ru"),
            ],
            [types.InlineKeyboardButton(text="🔗 Map linkni tahrirlash", callback_data="edit_map_link")],
            [types.InlineKeyboardButton(text="⬅️ Orqaga", callback_data="admin_back")],
        ]
    )


# --- Handlers ---
MAIN_MENU_BUTTONS = ("🏠 Asosiy menyu", "🏠 Главное меню")


@router.message(Command("admin"))
@admin_only
async def admin_panel(message: types.Message):
    lang = db.get_user_language(message.from_user.id)
    sticker_id = getattr(message.bot, "admin_sticker", "") or ""
    if sticker_id:
        try:
            await message.answer_sticker(sticker_id)
        except Exception:
            pass
    await message.answer("Admin panel" if lang == "uz" else "Админ-панель", reply_markup=admin_menu(lang))


@router.message(lambda m: m.text in MAIN_MENU_BUTTONS)
@admin_only
async def admin_to_main_menu(message: types.Message, state: FSMContext):
    """Return admin back to user main menu."""
    lang = db.get_user_language(message.from_user.id)
    await state.clear()
    prompt = "Asosiy menyu:" if lang == "uz" else "Главное меню:"
    await message.answer(prompt, reply_markup=main_menu(lang))


@router.message(lambda m: text_has(m, "treklar", "ro'yxati") or text_has(m, "список", "трек"))
@admin_only
async def list_tracks(message: types.Message):
    lang = db.get_user_language(message.from_user.id)
    tracks = db.list_tracks(limit=200)
    if not tracks:
        text = "Treklar topilmadi." if lang == "uz" else "Треков нет."
        await message.answer(text)
        return
    lines = []
    for row in tracks:
        lines.append(
            f"Trek: {row['track_code']}\nReys: {row['flight_number']}\nStatus: {row['status']}\n---"
        )
    if len(tracks) == 200:
        lines.append("... yana treklar bor." if lang == "uz" else "... есть еще треки.")
    await message.answer("\n".join(lines))


@router.message(lambda m: text_has(m, "elon") or text_has(m, "broadcast") or text_has(m, "рассылка") or text_has(m, "elon yuborish"))
@admin_only
async def broadcast_start(message: types.Message, state: FSMContext):
    lang = db.get_user_language(message.from_user.id)
    await state.set_state(BroadcastState.waiting_for_content)
    prompt = (
        "Matn, rasm, video yoki audio yuboring — barchaga jo'natamiz."
        if lang == "uz"
        else "Отправьте текст/фото/видео/аудио — разошлем всем."
    )
    await message.answer(prompt)


@router.message(BroadcastState.waiting_for_content)
@admin_only
async def broadcast_send(message: types.Message, state: FSMContext):
    lang = db.get_user_language(message.from_user.id)
    user_ids = db.list_user_ids()
    sent = 0
    failed = 0
    for uid in user_ids:
        try:
            await message.copy_to(uid)
            sent += 1
        except Exception:
            failed += 1
            continue
    await state.clear()
    text = f"Yuborildi: {sent}\nXato: {failed}" if lang == "uz" else f"Отправлено: {sent}\nОшибки: {failed}"
    await message.answer(text)


@router.message(lambda m: text_has(m, "trek kod qo'shish") or text_has(m, "trek", "qo'shish") or text_has(m, "добавить", "трек"))
@admin_only
async def add_track_start(message: types.Message, state: FSMContext):
    lang = db.get_user_language(message.from_user.id)
    await state.set_state(AddTrackState.waiting_for_flight)
    prompt = "Reys raqamini kiriting." if lang == "uz" else "Введите номер рейса."
    await message.answer(prompt)


@router.message(AddTrackState.waiting_for_flight)
@admin_only
async def add_track_get_flight(message: types.Message, state: FSMContext):
    await state.update_data(flight_number=message.text.strip())
    lang = db.get_user_language(message.from_user.id)
    prompt = "Trek kod(lar)ni kiriting." if lang == "uz" else "Введите трек-коды."
    await state.set_state(AddTrackState.waiting_for_codes)
    await message.answer(prompt)


@router.message(AddTrackState.waiting_for_codes)
@admin_only
async def add_track_save(message: types.Message, state: FSMContext):
    data = await state.get_data()
    flight = data.get("flight_number")
    codes = [code.strip() for code in message.text.replace(",", " ").split() if code.strip()]
    lang = db.get_user_language(message.from_user.id)
    default_status = "В пути" if lang == "ru" else "Yo'lda"
    for code in codes:
        db.add_track(code, flight, status=default_status)
    await state.clear()
    done = "Treklar saqlandi." if lang == "uz" else "Треки сохранены."
    await message.answer(done)


@router.message(lambda m: text_has(m, "trek kod o'chirish") or text_has(m, "удалить", "трек"))
@admin_only
async def delete_tracks_start(message: types.Message, state: FSMContext):
    lang = db.get_user_language(message.from_user.id)
    await state.set_state(DeleteTrackState.waiting_for_flight)
    prompt = "Reys raqamini kiriting." if lang == "uz" else "Введите номер рейса."
    await message.answer(prompt)


@router.message(DeleteTrackState.waiting_for_flight)
@admin_only
async def delete_tracks_by_flight(message: types.Message, state: FSMContext):
    lang = db.get_user_language(message.from_user.id)
    deleted = db.delete_tracks_by_flight(message.text.strip())
    await state.clear()
    if lang == "uz":
        text = "Ushbu reys bo'yicha trek kodlar o'chirildi." if deleted else "Hech narsa topilmadi."
    else:
        text = "Треки по рейсу удалены." if deleted else "Ничего не найдено."
    await message.answer(text)


@router.message(lambda m: text_has(m, "statistika") or text_has(m, "статистика"))
@admin_only
async def stats(message: types.Message):
    lang = db.get_user_language(message.from_user.id)
    stats_data = db.track_stats()
    users = db.user_count()
    admins = db.admin_count()
    text = (
        f"Foydalanuvchilar: {users}\n"
        f"Adminlar: {admins}\n"
        f"Jami treklar: {stats_data['total']}\n"
        f"Aktiv treklar: {stats_data['active']}\n"
        f"Oxirgi 24 soat: {stats_data['recent']}\n"
        f"So'nggi 7 kun: {stats_data['recent_7d']}"
        if lang == "uz"
        else f"Пользователей: {users}\n"
             f"Админов: {admins}\n"
             f"Всего треков: {stats_data['total']}\n"
             f"Активные: {stats_data['active']}\n"
             f"За 24 часа: {stats_data['recent']}\n"
             f"За 7 дней: {stats_data['recent_7d']}"
    )
    sticker_id = getattr(message.bot, "admin_sticker", "") or ""
    if sticker_id:
        try:
            await message.answer_sticker(sticker_id)
        except Exception:
            pass
    await message.answer(text)


@router.message(StateFilter(None), lambda m: text_has(m, "kanal ulash") or text_has(m, "канал"))
@admin_only
async def channel_start(message: types.Message, state: FSMContext):
    lang = db.get_user_language(message.from_user.id)
    await state.set_state(LinkChannelState.waiting_for_channel)
    current = db.get_setting("channel_username") or "sozlanmagan"
    prompt = (
        f"Hozirgi kanal: {current}\n"
        "Yangi kanal username yoki linkini yuboring (masalan, @mychannel yoki https://t.me/mychannel).\n"
        "Majburiy obunani o'chirish uchun: off"
        if lang == "uz"
        else f"Текущий канал: {current}\n"
             "Отправьте username или ссылку канала (например, @mychannel или https://t.me/mychannel).\n"
             "Чтобы отключить обязательную подписку: off"
    )
    await message.answer(prompt)


@router.message(LinkChannelState.waiting_for_channel)
@admin_only
async def channel_save(message: types.Message, state: FSMContext):
    value = (message.text or "").strip()
    await state.clear()
    lang = db.get_user_language(message.from_user.id)
    if value.lower() in {"off", "disable", "o'chirish", "ochirish", "0", "-"}:
        db.set_setting("channel_username", "")
        done = "Majburiy obuna o'chirildi." if lang == "uz" else "Обязательная подписка отключена."
        await message.answer(done, reply_markup=admin_menu(lang))
        return

    channel_link, channel_chat_id = normalize_channel(value)
    db.set_setting("channel_username", value)
    if channel_chat_id:
        done = (
            f"Kanal saqlandi: {channel_link}\n"
            "Eslatma: bot kanalga admin qilingan bo'lishi kerak, aks holda obunani tekshira olmaydi."
            if lang == "uz"
            else f"Канал сохранен: {channel_link}\n"
                 "Важно: бот должен быть админом канала, иначе он не сможет проверять подписку."
        )
    else:
        done = (
            f"Kanal linki saqlandi: {channel_link}\n"
            "Bu private invite linkga o'xshaydi. Telegram bunday link orqali obunani botdan tekshirtirmaydi; tekshiruv ishlashi uchun public @username yuboring."
            if lang == "uz"
            else f"Ссылка канала сохранена: {channel_link}\n"
                 "Это похоже на private invite link. Telegram не дает боту проверять подписку по такой ссылке; для проверки нужен public @username."
        )
    await message.answer(done, reply_markup=admin_menu(lang))


@router.message(lambda m: text_has(m, "admin qo'shish") or text_has(m, "добавить", "админ"))
@admin_only
async def add_admin_start(message: types.Message, state: FSMContext):
    lang = db.get_user_language(message.from_user.id)
    await state.set_state(AddAdminState.waiting_for_user)
    prompt = "Yangi adminning Telegram ID yoki username'ini yuboring." if lang == "uz" else "Отправьте Telegram ID или username нового админа."
    await message.answer(prompt)


@router.message(AddAdminState.waiting_for_user)
@admin_only
async def add_admin_save(message: types.Message, state: FSMContext):
    value = message.text.strip()
    lang = db.get_user_language(message.from_user.id)
    if value.startswith("@"):
        text = "Iltimos, raqamli Telegram ID yuboring." if lang == "uz" else "Отправьте числовой Telegram ID."
        await message.answer(text)
        return
    try:
        user_id = int(value)
    except ValueError:
        await message.answer("ID noto'g'ri." if lang == "uz" else "Некорректный ID.")
        return
    db.add_admin(user_id)
    await state.clear()
    await message.answer("Admin qo'shildi." if lang == "uz" else "Админ добавлен.")


# --- About edit ---
ABOUT_BUTTONS = (
    "ℹ️ Biz haqimizda (edit)",
    "ℹ️ О нас (ред.)",
    "Biz haqimizda (edit)",
)
PRICE_BUTTONS = (
    "💰 Narxlar (edit)",
    "💰 Цены (ред.)",
    "Narxlar (edit)",
)
ADDRESS_BUTTONS = (
    "📍 Manzil (edit)",
    "📍 Адрес (ред.)",
    "Manzil (edit)",
)


@router.message(lambda m: m.text in ABOUT_BUTTONS or (m.text and (text_has(m, "haqimizda") and text_has(m, "edit")) or (text_has(m, "нас") and text_has(m, "ред"))))
@admin_only
async def edit_about_start(message: types.Message, state: FSMContext):
    lang = db.get_user_language(message.from_user.id)
    about_uz = db.get_setting("about_uz") or ""
    about_ru = db.get_setting("about_ru") or ""
    await message.answer(
        f"Hozirgi (UZ):\\n{about_uz}\\n\\nHozirgi (RU):\\n{about_ru}",
        reply_markup=about_edit_keyboard(),
    )
    await state.set_state(AboutEditState.choosing)
    prompt = "O'zgartirmoqchi bo'lgan matnni tanlang." if lang == "uz" else "Выберите, что хотите изменить."
    await message.answer(prompt)


@router.callback_query(lambda c: c.data == "edit_about_uz")
@admin_only_callback
async def edit_about_choose_uz(callback: types.CallbackQuery, state: FSMContext):
    await state.set_state(AboutEditState.waiting_for_uz_text)
    await callback.message.answer("Yangi o'zbekcha 'Biz haqimizda' matnini yuboring.")
    await callback.answer()


@router.callback_query(lambda c: c.data == "edit_about_ru")
@admin_only_callback
async def edit_about_choose_ru(callback: types.CallbackQuery, state: FSMContext):
    await state.set_state(AboutEditState.waiting_for_ru_text)
    await callback.message.answer("Отправьте новый текст 'О нас' на русском.")
    await callback.answer()


@router.callback_query(lambda c: c.data == "admin_back")
@admin_only_callback
async def admin_back(callback: types.CallbackQuery, state: FSMContext):
    lang = db.get_user_language(callback.from_user.id)
    await state.clear()
    await callback.message.answer(
        "Admin panel" if lang == "uz" else "Админ-панель",
        reply_markup=admin_menu(lang),
    )
    await callback.answer()


@router.message(AboutEditState.waiting_for_uz_text)
@admin_only
async def edit_about_uz(message: types.Message, state: FSMContext):
    lang = db.get_user_language(message.from_user.id)
    db.set_setting("about_uz", message.text)
    await state.clear()
    confirm = "O'zbekcha matn yangilandi." if lang == "uz" else "Текст на узбекском обновлен."
    await message.answer(confirm, reply_markup=admin_menu(lang))


@router.message(AboutEditState.waiting_for_ru_text)
@admin_only
async def edit_about_ru(message: types.Message, state: FSMContext):
    lang = db.get_user_language(message.from_user.id)
    db.set_setting("about_ru", message.text)
    await state.clear()
    confirm = "Ruscha matn yangilandi." if lang == "uz" else "Русский текст обновлён."
    await message.answer(confirm, reply_markup=admin_menu(lang))




# --- Price edit ---
@router.message(lambda m: m.text in PRICE_BUTTONS or (m.text and (text_has(m, "narxlar") and text_has(m, "edit")) or (text_has(m, "цены") and text_has(m, "ред"))))
@admin_only
async def edit_price_start(message: types.Message, state: FSMContext):
    lang = db.get_user_language(message.from_user.id)
    price_uz = db.get_setting("prices_uz") or db.get_setting("price_uz") or ""
    price_ru = db.get_setting("prices_ru") or db.get_setting("price_ru") or ""
    await message.answer(
        f"Hozirgi (UZ):\\n{price_uz}\\n\\nHozirgi (RU):\\n{price_ru}",
        reply_markup=price_edit_keyboard(),
    )
    await state.set_state(PriceEditState.choosing)
    prompt = "Qaysi narxni tahrirlaysiz?" if lang == "uz" else "Что редактируем?"
    await message.answer(prompt)


@router.callback_query(lambda c: c.data == "edit_prices_uz")
@admin_only_callback
async def edit_price_choose_uz(callback: types.CallbackQuery, state: FSMContext):
    await state.set_state(PriceEditState.waiting_for_uz_text)
    await callback.message.answer(
        "Yangi o‘zbekcha narxlar matnini yuboring (masalan: 'Avto: 6$ / kg\\nAvia: 9$ / kg')."
    )
    await callback.answer()


@router.callback_query(lambda c: c.data == "edit_prices_ru")
@admin_only_callback
async def edit_price_choose_ru(callback: types.CallbackQuery, state: FSMContext):
    await state.set_state(PriceEditState.waiting_for_ru_text)
    await callback.message.answer("Отправьте новый текст цен на русском.")
    await callback.answer()


@router.message(PriceEditState.waiting_for_uz_text)
@admin_only
async def edit_price_save_uz(message: types.Message, state: FSMContext):
    lang = db.get_user_language(message.from_user.id)
    db.set_setting("prices_uz", message.text)
    db.set_setting("price_uz", message.text)  # keep compatibility
    await state.clear()
    confirm = "O‘zbekcha narxlar muvaffaqiyatli yangilandi." if lang == "uz" else "Цены на узбекском обновлены."
    await message.answer(confirm, reply_markup=admin_menu(lang))


@router.message(PriceEditState.waiting_for_ru_text)
@admin_only
async def edit_price_save_ru(message: types.Message, state: FSMContext):
    lang = db.get_user_language(message.from_user.id)
    db.set_setting("prices_ru", message.text)
    db.set_setting("price_ru", message.text)  # keep compatibility
    await state.clear()
    confirm = "Ruscha narxlar muvaffaqiyatli yangilandi." if lang == "uz" else "Русские цены успешно обновлены."
    await message.answer(confirm, reply_markup=admin_menu(lang))




# --- Address edit ---
@router.message(lambda m: m.text in ADDRESS_BUTTONS or (m.text and (text_has(m, "manzil") and text_has(m, "edit")) or (text_has(m, "адрес") and text_has(m, "ред"))))
@admin_only
async def edit_address_start(message: types.Message, state: FSMContext):
    lang = db.get_user_language(message.from_user.id)
    address_uz = db.get_setting("address_uz") or ""
    address_ru = db.get_setting("address_ru") or ""
    map_link = db.get_setting("map_link") or ""
    await message.answer(
        f"Hozirgi manzil (UZ):\\n{address_uz}\\n\\n(RU):\\n{address_ru}\\n\\nMap: {map_link}",
        reply_markup=address_edit_keyboard(),
    )
    await state.set_state(AddressEditState.choosing)
    prompt = "Qaysi qismini tahrirlaysiz?" if lang == "uz" else "Что редактируем?"
    await message.answer(prompt)


@router.callback_query(lambda c: c.data == "edit_address_uz")
@admin_only_callback
async def edit_address_choose_uz(callback: types.CallbackQuery, state: FSMContext):
    await state.set_state(AddressEditState.waiting_for_uz_text)
    await callback.message.answer("Yangi o'zbekcha manzil matnini yuboring.")
    await callback.answer()


@router.callback_query(lambda c: c.data == "edit_address_ru")
@admin_only_callback
async def edit_address_choose_ru(callback: types.CallbackQuery, state: FSMContext):
    await state.set_state(AddressEditState.waiting_for_ru_text)
    await callback.message.answer("Отправьте новый текст адреса на русском.")
    await callback.answer()


@router.callback_query(lambda c: c.data == "edit_map_link")
@admin_only_callback
async def edit_address_choose_map(callback: types.CallbackQuery, state: FSMContext):
    await state.set_state(AddressEditState.waiting_for_map_link)
    await callback.message.answer("Yangi Google Maps linkini yuboring (yoki lokatsiya).")
    await callback.answer()


@router.message(AddressEditState.waiting_for_uz_text)
@admin_only
async def edit_address_save(message: types.Message, state: FSMContext):
    lang = db.get_user_language(message.from_user.id)
    db.set_setting("address_uz", message.text)
    await state.clear()
    confirm = "O'zbekcha manzil yangilandi." if lang == "uz" else "Адрес на узбекском обновлен."
    await message.answer(confirm, reply_markup=admin_menu(lang))


@router.message(AddressEditState.waiting_for_ru_text)
@admin_only
async def edit_address_save_ru(message: types.Message, state: FSMContext):
    lang = db.get_user_language(message.from_user.id)
    db.set_setting("address_ru", message.text)
    await state.clear()
    confirm = "Ruscha manzil yangilandi." if lang == "uz" else "Адрес на русском обновлен."
    await message.answer(confirm, reply_markup=admin_menu(lang))


@router.message(AddressEditState.waiting_for_map_link)
@admin_only
async def edit_address_save_map(message: types.Message, state: FSMContext):
    lang = db.get_user_language(message.from_user.id)
    db.set_setting("map_link", message.text)
    await state.clear()
    confirm = "Map link yangilandi." if lang == "uz" else "Ссылка на карту обновлена."
    await message.answer(confirm, reply_markup=admin_menu(lang))


# --- Excel import ---
@router.message(lambda m: text_has(m, "import") and text_has(m, "trek") or text_has(m, "импорт") and text_has(m, "трек"))
@admin_only
async def import_tracks_start(message: types.Message, state: FSMContext):
    lang = db.get_user_language(message.from_user.id)
    await state.set_state(ImportTracksState.waiting_for_file)
    prompt = "Excel faylni yuboring (xlsx)." if lang == "uz" else "Отправьте Excel файл (xlsx)."
    await message.answer(prompt)


async def _finalize_import(message: types.Message, entries, lang: str, default_status: str, missing_track_rows: int = 0):
    added = 0
    skipped = missing_track_rows
    duplicates = 0
    missing_data = missing_track_rows

    for entry in entries:
        track_code = (entry.get("track") or "").strip()
        flight_number = (entry.get("flight") or "").strip()
        status_value = entry.get("status") or default_status
        if not track_code or not flight_number:
            skipped += 1
            missing_data += 1
            continue
        if db.get_track_by_code(track_code):
            skipped += 1
            duplicates += 1
            continue
        try:
            inserted = db.add_track(track_code, flight_number, status_value or default_status)
            if inserted:
                added += 1
            else:
                skipped += 1
        except Exception:
            skipped += 1

    summary_lines = []
    if lang == "uz":
        summary_lines.append("Import yakunlandi.")
        summary_lines.append(f"Qo'shildi: {added}")
        summary_lines.append(f"O'tkazib yuborildi/xato: {skipped}")
        if duplicates:
            summary_lines.append(f"- Dublikat treklar: {duplicates}")
        if missing_data:
            summary_lines.append(f"- Bo'sh ustunlar: {missing_data}")
        if not added and not skipped:
            summary_lines.append("Faylda satrlar topilmadi.")
    else:
        summary_lines.append("Импорт завершен.")
        summary_lines.append(f"Добавлено: {added}")
        summary_lines.append(f"Пропущено/ошибка: {skipped}")
        if duplicates:
            summary_lines.append(f"- Дубликаты треков: {duplicates}")
        if missing_data:
            summary_lines.append(f"- Пустые столбцы: {missing_data}")
        if not added and not skipped:
            summary_lines.append("Строки в файле не найдены.")
    await message.answer("\n".join(summary_lines))


@router.message(ImportTracksState.waiting_for_file)
@admin_only
async def import_tracks_file(message: types.Message, state: FSMContext):
    lang = db.get_user_language(message.from_user.id)
    if not message.document or not message.document.file_name.lower().endswith(".xlsx"):
        await message.answer("Faqat .xlsx fayl yuboring." if lang == "uz" else "Отправьте файл в формате .xlsx.")
        return

    from openpyxl import load_workbook

    default_status = "В пути" if lang == "ru" else "Yo'lda"
    temp = NamedTemporaryFile(delete=False, suffix=".xlsx")
    temp_path = temp.name
    temp.close()
    try:
        await message.bot.download(file=message.document, destination=temp_path)
    except Exception as exc:  # noqa: BLE001
        logging.error("Excel import: failed to download file: %s", exc)
        text = (
            "Faylni yuklab bo'lmadi, qaytadan yuboring."
            if lang == "uz"
            else "Не удалось скачать файл, отправьте заново."
        )
        await message.answer(text)
        return

    entries = []
    missing_track_rows = 0
    row_index = 0
    try:
        wb = load_workbook(temp_path, read_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=1, values_only=True):
            row_index += 1
            if not row:
                continue
            track_code = str(row[0]).strip() if row[0] else ""
            flight_number = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            status_value = str(row[2]).strip() if len(row) > 2 and row[2] else default_status
            # Skip header row that looks like titles
            if row_index == 1:
                header_text = f"{track_code} {flight_number} {status_value}".lower()
                if any(key in header_text for key in ("track", "trek", "трек", "код", "рейс", "status", "статус")):
                    continue
            if not track_code:
                missing_track_rows += 1
                continue
            entries.append({"track": track_code, "flight": flight_number, "status": status_value})
    except Exception as exc:  # noqa: BLE001
        logging.error("Excel import: failed to process workbook: %s", exc)
        text = (
            "Faylni o‘qib bo‘lmadi, formatni tekshiring."
            if lang == "uz"
            else "Не удалось прочитать файл, проверьте формат."
        )
        await message.answer(text)
        return
    finally:
        try:
            os.remove(temp_path)
        except Exception:
            pass

    if not entries:
        text = "Faylda trek kodlar topilmadi." if lang == "uz" else "Трек-коды в файле не найдены."
        await message.answer(text)
        await state.clear()
        return

    need_flight = any(not e.get("flight") for e in entries)
    if need_flight:
        await state.update_data(pending_import=entries, missing_track_rows=missing_track_rows, default_status=default_status)
        await state.set_state(ImportTracksState.waiting_for_flight)
        prompt = (
            "Reys raqamini kiriting (faylda ko'rsatilmagan)."
            if lang == "uz"
            else "Введите номер рейса (в файле не указан)."
        )
        await message.answer(prompt)
        return

    await _finalize_import(message, entries, lang, default_status, missing_track_rows)
    await state.clear()


@router.message(ImportTracksState.waiting_for_flight)
@admin_only
async def import_tracks_set_flight(message: types.Message, state: FSMContext):
    lang = db.get_user_language(message.from_user.id)
    data = await state.get_data()
    entries = data.get("pending_import") or []
    default_status = data.get("default_status") or ("В пути" if lang == "ru" else "Yo'lda")
    missing_track_rows = data.get("missing_track_rows") or 0
    flight_number = message.text.strip()
    for entry in entries:
        if not entry.get("flight"):
            entry["flight"] = flight_number
    await _finalize_import(message, entries, lang, default_status, missing_track_rows)
    await state.clear()
